import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import cm
import matplotlib.colors as colors
import matplotlib as mpl
from matplotlib.ticker import FuncFormatter

import numpy as np
from scipy.interpolate import Rbf
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import load_workbook
import matplotlib.patches as mpatches

from matplotlib.cm import ScalarMappable
from matplotlib.colorbar import ColorbarBase

def format_z_3d(value, tick_number):
    # 将最高处的刻度更改为'>1000'
    if value == 1000:
        return '> 1000'
    else:
        return int(value)




wb_sheet_final = load_workbook('F:/chen_han_paper-python-graph/medicalPlotAgain/original_27_summary1.xlsx')
# 获取所有工作表的名称
sheet_names = wb_sheet_final.sheetnames
print(sheet_names)
esp_pic_name=sheet_names[4]+'_'

# 读取Excel文件
wb = load_workbook('F:/chen_han_paper-python-graph/medicalPlotAgain/data_temp.xlsx')
ws = wb.active

# 初始化列表
data = []

# 遍历所有行和列,读取单元格的值
for row in ws.iter_rows():
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    data.append(row_data)

# 使用pandas转换为Dataframe
df = pd.DataFrame(data)


########################################################
########################################################
########################################################

## 设置 dpi
mpl.rcParams['figure.dpi'] = 300

## 以下二维
fig, ax = plt.subplots()

maxium_num = -100
minium_num = 100

Z_matrix = np.zeros((36,96))
X_axis = ['']*96
Y_axis = ['']*36
# 绘制热力图
for iCnt in range(0,96): 
    for jCnt in range(0,36):
        if(df[jCnt+1][iCnt+1] == '#DIV/0!'):
            continue
        Z_matrix[jCnt][iCnt] = df[jCnt+1][96-iCnt]
        ##format_Z_matrix = "Z_matrix[{}][{}] = {}".format(jCnt,iCnt,Z_matrix[jCnt][iCnt])
        ##print(format_Z_matrix)
        if(Z_matrix[jCnt][iCnt] > maxium_num):
            maxium_num = Z_matrix[jCnt][iCnt]
        if(Z_matrix[jCnt][iCnt] < minium_num):
            minium_num = Z_matrix[jCnt][iCnt]

        X_axis[iCnt] = df[0][96-iCnt]
        Y_axis[jCnt] = df[jCnt+1][0]
        ## print("x,y,z:",df[0][iCnt+1],df[jCnt+1][0],df[jCnt+1][iCnt+1])
 
print("X_axis  detail:",X_axis)
print("Y_axis  detail:",Y_axis)
print("maxium_num:",maxium_num)
print("minium_num:",minium_num)

## 将Z_matrix值范围缩放到0-1之间:
## 便于后续映射到热力图
Z_scaled  = (Z_matrix - minium_num) / (maxium_num - minium_num)


# 应用到某个已有的颜色条cmap上
# new_cmap = cmap_map(function, cm.cool)
#分段定义红白色调色条
own_designer_cmap = plt.cm.get_cmap('jet_r')

##boundaries = np.linspace(-0.95, 0.30, 257)
boundaries = np.concatenate((np.linspace(0,1,128),np.linspace(1,0,128)))

## norm = cm.colors.BoundaryNorm(boundaries, own_designer_cmap.N, clip=True)
## norm = cm.colors.Normalize()
norm = cm.colors.TwoSlopeNorm(vmin=-1.125, vcenter=0, vmax=0.375)


bDrawHeatMap = False
bDrawGridLine = True
bDrawContour = False
if bDrawHeatMap:
    im = ax.imshow(Z_matrix,own_designer_cmap,interpolation='nearest',origin='lower',extent=[0,96,0,41],aspect='auto',norm=norm)

    nn5 = 13  # 指定n的值
    array_colorBar_tick = np.linspace(-1.125, 0.375, nn5)
    array_colorBar_tick = np.around(array_colorBar_tick, 3)  # 将数组元素四舍五入到两位小数
    #手工填入定制刻度
    ##array_colorBar_tick = [-1.25,-1.125,-1.0,-0.875,-0.75,-0.625,-0.5,-0.375,-0.25, -0.125, 0.0 ,0.125,0.25,0.375,0.5,0.625]
    cbar = fig.colorbar(im,ticks=array_colorBar_tick)
    # 在最顶部和最底部添加正无穷和负无穷的值,并且过滤掉（-1，0.5）之外的值
    curr_labels = [item.get_text() for item in cbar.ax.get_yticklabels()]
    ##curr_labels[0] = '<-1.0'
    ##curr_labels[1] = ''
    ##curr_labels[-1] = '>0.5'

    cbar.ax.set_yticklabels(curr_labels, fontsize=6)
    cbar.ax.invert_yaxis()

    esp_pic_name=esp_pic_name+"heatmap"

    if bDrawContour:
        # 绘制等高线      levels=(-0.5,-0.2,0)
        cs_same_height = ax.contour(Z_matrix, np.linspace(minium_num,maxium_num,6), colors='k', linestyles='--', linewidths=0.3)
        # 绘制等高线 数值说明
        plt.clabel(cs_same_height, inline=True, fontsize=5)
        esp_pic_name=esp_pic_name+"等高线"
elif bDrawGridLine:
    # 绘制gridline
    plt.grid(True, linestyle='--', linewidth=0.5, alpha=0.5)
    esp_pic_name=esp_pic_name+"gridline"

   

# 设置图形标题  
# ax.set_title('2D Scatter Plot Medical Heatmap',fontsize=25,fontweight='bold')
 
# 设置坐标轴标签
x_ticks = np.arange(0, 97, 4)
x_tick_labels = [f"{i/2}h" for i in np.arange(0, 97, 4)] 
ax.set_xticks(x_ticks)
ax.set_xticklabels(x_tick_labels,fontsize=4)
ax.set_xlabel('Time (hours)',fontsize=8,fontweight='bold')

y_ticks = np.arange(5, 41, 5)
y_tick_labels = [f"{i}" for i in np.arange(5, 41, 5)]
ax.set_yticks(y_ticks)  
ax.set_yticklabels(y_tick_labels,fontsize=6)
ax.set_ylabel('Intracranial Pressure (mmHg)',fontsize=10,fontweight='bold')


########################################################
########################################################
########################################################

bDrawThrehold = True
if bDrawThrehold:
    esp_pic_name=esp_pic_name+"threshold"

# 设置阈值
threshold_1 = -1.0

# 根据阈值绘制曲线
x_curve = []
y_curve = []
for i in range(Z_matrix.shape[1]):
    for j in range(Z_matrix.shape[0]):
        if Z_matrix[j, i] <= threshold_1 + 0.05 and Z_matrix[j, i] >= threshold_1 - 0.05:
            x_curve.append(i + 1)
            y_curve.append(j + 5)
            break  # 只选取每列的第一个满足条件的点

y_curve[34] = 23
y_curve[35] = 23
y_curve[36] = 22
y_curve[37] = 23
y_curve[38] = 23
y_curve[39] = 23
y_curve[40] = 23
y_curve[41] = 23
y_curve[42] = 22
y_curve[43] = 22
y_curve[44] = 22
y_curve[45] = 21
y_curve[46] = 22
y_curve[47] = 22
y_curve[48] = 22
y_curve[49] = 21
y_curve[50] = 21
y_curve[51] = 21
y_curve[52] = 21
y_curve[53] = 21
y_curve[54] = 21
y_curve[55] = 21
y_curve[56] = 21
y_curve[57] = 21
y_curve[58] = 21
y_curve[59] = 21
y_curve[60] = 21
y_curve[61] = 21
y_curve[62] = 21
y_curve[63] = 21
y_curve[64] = 21
y_curve[65] = 21
y_curve[66] = 21
y_curve[67] = 21
y_curve[68] = 21
y_curve[69] = 21
y_curve[70] = 21
y_curve[71] = 21
y_curve[72] = 21
y_curve[73] = 21
y_curve[74] = 21
y_curve[75] = 21
y_curve[76] = 21
y_curve[77] = 21
y_curve[78] = 21
y_curve[79] = 21
y_curve[80] = 21
y_curve[81] = 21
y_curve[82] = 21
y_curve[83] = 21
y_curve[84] = 21
y_curve[85] = 21
y_curve[86] = 21
y_curve[87] = 20
y_curve[88] = 20
y_curve[89] = 20
y_curve[90] = 20
y_curve[91] = 20
y_curve[92] = 20
y_curve[93] = 20
print("zaojia 1 x_curve: ", x_curve)
print("zaojia 1 y_curve: ", y_curve)

if bDrawThrehold:
    ax.plot(x_curve, y_curve, color='darkgray', linewidth=1,)

# x_end = x_curve[0]  
# y_end = y_curve[0]
# plt.annotate('threshold:  0 +- 0.05) ', xy=(x_end, y_end), xytext=(x_end-15, y_end+1), 
#             arrowprops=dict(arrowstyle='->', connectionstyle='arc3', color='gray'))

# 设置阈值
threshold_2 = -0.5
x_curve = []
y_curve = []

# 根据阈值绘制曲线
for i in range(Z_matrix.shape[1]):
    for j in range(Z_matrix.shape[0]):
        if Z_matrix[j, i] <= threshold_2 + 0.02 and Z_matrix[j, i] >= threshold_2 - 0.02:
            x_curve.append(i + 1)
            y_curve.append(j + 5)
            break  # 只选取每列的第一个满足条件的点

if bDrawThrehold:
    ax.plot(x_curve, y_curve, color='purple', linewidth=1,)



# 设置阈值
threshold_3 = 0.0
x_curve = []
y_curve = []

# 根据阈值绘制曲线
for i in range(Z_matrix.shape[1]):
    for j in range(Z_matrix.shape[0]):
        if Z_matrix[j, i] <= threshold_3 + 0.03 and Z_matrix[j, i] >= threshold_3 - 0.03:
            x_curve.append(i + 1)
            y_curve.append(j + 5)
            break  # 只选取每列的第一个满足条件的点

if bDrawThrehold:
    ax.plot(x_curve, y_curve, color='black', linewidth=1 ,)

# x_end = x_curve[0]  
# y_end = y_curve[0]
# plt.annotate('threshold:  -0.2 +- 0.05) ', xy=(x_end, y_end), xytext=(x_end-20, y_end+1), 
#             arrowprops=dict(arrowstyle='->', connectionstyle='arc3', color='gray'))

bSavePicAndShow = True
if bSavePicAndShow:
    plt.savefig( esp_pic_name+'.eps', format='eps',dpi=300)
    plt.savefig( esp_pic_name+'.png', dpi=300)      ## esp_pic_name+
    plt.show()

# ########################################################
# ########################################################
# ########################################################

# # 以下三维
# _X, _Y = np.meshgrid(X_axis, Y_axis)
# X_ravel = _X.ravel()
# Y_ravel = _Y.ravel()
# Z_ravel = np.zeros_like(X_ravel)
# dx = dy = 1
# dz = Z_matrix.ravel()

# # # 创建三维图 https://pypi.tuna.tsinghua.edu.cn/simple
# fig_3d = plt.figure()
# ax3D = plt.axes(projection='3d')


# value_ranges = [(0, 10), (10, 20), (20, 30), (30, 40), (40, 50), (50, 60), (60, 70), (70, 80), (80, 90), (90, 100), 
#                 (100, 110), (110, 120), (120, 130), (130, 140), (140, 150), (150, 200), (200, 300), (300, 400), (400, 500), (500, 600), (600, 700),(700, 800), (800, 900),(900, 1000)]
#                 # ,(2000, 3000),(3000, 4000),(4000, 5000),(5000, 6000),(6000, 7000),(7000, 8000),(8000, 9000),(9000, 10000),(10000, 11000),(11000, 12000),(12000, 13000)]
# color_map_3d = cm.get_cmap('cool', len(value_ranges) - 1)
# norm_3d = colors.Normalize(vmin=minium_num, vmax=maxium_num)

# color_arr = []
# for val in dz:
#     found = False
#     for range_min, range_max in value_ranges:
#         if range_min <= val <= range_max:
#             color_arr.append(color_map_3d(norm_3d(val)))
#             found = True
#             break
#     if not found:
#         color_arr.append(None)
# ax3D.bar3d(X_ravel, Y_ravel, Z_ravel, dx, dy, dz, color=color_arr)

# # Plot the 3D surface
# ##surf = ax3D.plot_surface(X_axis, Y_axis, Z_matrix, cmap=own_designer_cmap, linewidth=0, antialiased=False)

# # Set the limits of the z axis, so the color bar will match this range
# # 设置坐标轴范围
# ##ax3D.set_zlim(minium_num, maxium_num)

# # Set color bar and stipulate its size
# ##fig.colorbar(surf, shrink=0.5, aspect=10)

# # Set title and axis labels
# ax3D.set_title('Survivor', fontsize=8, fontweight='bold')
# ax3D.set_xlabel('Time (hours)', fontsize=6, fontweight='bold')
# ax3D.set_ylabel('Intracranial Pressure (mmHg)', fontsize=6, fontweight='bold')


# # 创建FuncFormatter  1000以上的值都format一下
# formatter = FuncFormatter(format_z_3d)
# # 将自定义的刻度格式应用到z轴
# ax3D.zaxis.set_major_formatter(formatter)
# ax3D.set_zlabel('Count', fontsize=6, fontweight='bold')


# ax3D.view_init(elev=10, azim=50) # 调整视角
# ##ax3D.dist = 7 # 设置相机距离


# #最大化显示图形窗口
# figManager = plt.get_current_fig_manager() 
# figManager.resize(*figManager.window.maxsize())

# plt.savefig('Survivor_3DScatterPlotMedicalData.png', dpi=300)
# plt.savefig('Survivor_3DScatterPlotMedicalData.eps', format='eps',dpi=300)

# plt.show()
